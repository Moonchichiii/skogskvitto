import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from receipts.models import Receipt


def build_excel(receipts: list[Receipt]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Kvitton"

    headers = ["Datum", "Företag", "Kategori", "Netto", "Moms", "Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    col_widths = [12, 28, 18, 12, 12, 12]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, receipt in enumerate(receipts, start=2):
        total = receipt.total_amount or Decimal(0)
        vat = receipt.vat_amount or Decimal(0)
        net = total - vat
        has_amount = receipt.total_amount is not None or receipt.vat_amount is not None
        ws.cell(row=row_idx, column=1, value=str(receipt.date) if receipt.date else "")
        ws.cell(row=row_idx, column=2, value=receipt.vendor or "")
        ws.cell(row=row_idx, column=3, value=receipt.category or "")
        ws.cell(row=row_idx, column=4, value=float(net) if has_amount else "")
        ws.cell(row=row_idx, column=5, value=float(vat) if receipt.vat_amount else "")
        ws.cell(row=row_idx, column=6, value=float(total) if receipt.total_amount else "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
