import io
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apps.receipts.models import Receipt

ZERO = Decimal("0.00")
SKOGSAVDRAGS_MOMS_RATE = Decimal("0.25")
KORJOURNAL_RATE_PER_MIL = Decimal("25")
KM_PER_MIL = Decimal("10")
CURRENCY_FORMAT = '#,##0.00 "kr"'
NUMBER_FORMAT = "#,##0.00"


@dataclass
class MonthlyDrive:
    trips: int = 0
    kilometers: Decimal = ZERO


def _decimal(value: Decimal | None) -> Decimal:
    return value if value is not None else ZERO


def _normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return normalized.encode("ascii", "ignore").decode("ascii").lower()


def _setup_sheet(ws: Worksheet, headers: list[str], widths: list[int]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"


def _write_kvitton_sheet(wb: Workbook, receipts: list[Receipt]) -> None:
    ws = wb.active
    assert ws is not None
    ws.title = "Kvitton"

    headers = ["inköp", "netto", "moms", "öresutjämning"]
    _setup_sheet(ws, headers, [16, 16, 16, 18])

    total_inkop = ZERO
    total_netto = ZERO
    total_moms = ZERO
    total_oresutjamning = ZERO

    row = 2
    for receipt in receipts:
        has_amount = receipt.total_amount is not None or receipt.vat_amount is not None
        if not has_amount:
            continue

        inkop = _decimal(receipt.total_amount)
        moms = _decimal(receipt.vat_amount)
        netto = inkop - moms
        # TODO: Koppla öresutjämning till en explicit modellkälla när fält finns.
        oresutjamning = ZERO

        ws.cell(row=row, column=1, value=inkop)
        ws.cell(row=row, column=2, value=netto)
        ws.cell(row=row, column=3, value=moms)
        ws.cell(row=row, column=4, value=oresutjamning)

        total_inkop += inkop
        total_netto += netto
        total_moms += moms
        total_oresutjamning += oresutjamning
        row += 1

    total_row = row
    for col, value in enumerate(
        [total_inkop, total_netto, total_moms, total_oresutjamning],
        start=1,
    ):
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(bold=True)

    for col in range(1, 5):
        for row_idx in range(2, total_row + 1):
            ws.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT


def _write_inkomster_skog_sheet(wb: Workbook, receipts: list[Receipt]) -> None:
    ws = wb.create_sheet("Inkomster Skog")
    headers = ["Slutavverkning", "Gallring", "Flis", "Skogsavdragsmoms (25%)"]
    _setup_sheet(ws, headers, [20, 14, 12, 24])

    inkomster = {
        "Slutavverkning": ZERO,
        "Gallring": ZERO,
        "Flis": ZERO,
    }

    for receipt in receipts:
        amount = _decimal(receipt.total_amount)
        if amount == ZERO:
            continue

        category = _normalize_text((receipt.category or "").strip())
        if "slutavverk" in category:
            inkomster["Slutavverkning"] += amount
        elif "gallring" in category:
            inkomster["Gallring"] += amount
        elif "flis" in category:
            inkomster["Flis"] += amount

    skogsavdragsmoms = (
        (inkomster["Slutavverkning"] + inkomster["Gallring"] + inkomster["Flis"])
        * SKOGSAVDRAGS_MOMS_RATE
    )

    ws.cell(row=2, column=1, value=inkomster["Slutavverkning"])
    ws.cell(row=2, column=2, value=inkomster["Gallring"])
    ws.cell(row=2, column=3, value=inkomster["Flis"])
    total_moms_cell = ws.cell(row=2, column=4, value=skogsavdragsmoms)
    total_moms_cell.font = Font(bold=True)

    for col in range(1, 5):
        ws.cell(row=2, column=col).number_format = CURRENCY_FORMAT


def _write_korjournal_sheet(wb: Workbook, receipts: list[Receipt]) -> None:
    ws = wb.create_sheet("Körjournal")
    headers = ["Månad", "Antal resor", "Kilometer totalt", "Ersättning (25 kr/mil)"]
    _setup_sheet(ws, headers, [14, 14, 18, 22])

    monthly: defaultdict[str, MonthlyDrive] = defaultdict(MonthlyDrive)

    for receipt in receipts:
        if not receipt.date:
            continue

        normalized_category = _normalize_text((receipt.category or "").strip())
        if "korjournal" not in normalized_category:
            continue

        month_key = receipt.date.strftime("%Y-%m")
        month_data = monthly[month_key]
        month_data.trips += 1
        # TODO: Körjournal bör få dedikerat kilometerfält i modellen i stället för total_amount.
        month_data.kilometers += _decimal(receipt.total_amount)

    total_trips = 0
    total_kilometers = ZERO
    total_ersattning = ZERO

    row = 2
    for month in sorted(monthly.keys()):
        month_data = monthly[month]
        trips = month_data.trips
        kilometers = month_data.kilometers
        ersattning = (kilometers / KM_PER_MIL) * KORJOURNAL_RATE_PER_MIL

        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=trips)
        ws.cell(row=row, column=3, value=kilometers)
        ws.cell(row=row, column=4, value=ersattning)

        total_trips += trips
        total_kilometers += kilometers
        total_ersattning += ersattning
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="Summa").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=total_trips).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=total_kilometers).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total_ersattning).font = Font(bold=True)

    for row_idx in range(2, total_row + 1):
        ws.cell(row=row_idx, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=row_idx, column=4).number_format = CURRENCY_FORMAT


def build_excel(receipts: list[Receipt]) -> io.BytesIO:
    wb = Workbook()
    _write_kvitton_sheet(wb, receipts)
    _write_inkomster_skog_sheet(wb, receipts)
    _write_korjournal_sheet(wb, receipts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
