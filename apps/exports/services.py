"""Excel export — turns a list of receipts into a polished .xlsx workbook.

The export produces two sheets:
  - "Kvitton"        Full transaction list, one row per receipt
  - "Sammanfattning" Per-category sums + grand total

The "Kvitton" sheet is what the accountant reads. The "Sammanfattning" sheet
is for the owner's own overview.
"""

from __future__ import annotations

import io
from collections import OrderedDict
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apps.receipts.models import Receipt

# --- Styling tokens -----------------------------------------------------------

ZERO = Decimal("0.00")
CURRENCY_FORMAT = '#,##0.00 "kr"'

FOREST_DEEP = "FF00291F"  # ARGB — dark green header background
FOREST_SOFT = "FFE7F2EA"  # light green for total rows
STONE_200 = "FFE7E5E4"    # subtle row separator

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=FOREST_DEEP)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

BODY_FONT = Font(name="Calibri", size=11)
BODY_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
BODY_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
TOTAL_FILL = PatternFill("solid", fgColor=FOREST_SOFT)

BORDER = Border(bottom=Side(style="thin", color=STONE_200))


# --- Helpers ------------------------------------------------------------------


def _set_column_widths(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[row].height = 22


def _category_label(receipt: Receipt) -> str:
    """Normalised category for grouping — empty values become 'Okategoriserat'."""
    raw = (receipt.category or "").strip()
    return raw or "Okategoriserat"


# --- Sheet 1: Kvitton (full transaction list) --------------------------------


def _write_kvitton_sheet(wb: Workbook, receipts: list[Receipt]) -> None:
    ws = wb.active
    assert ws is not None
    ws.title = "Kvitton"

    headers = ["Datum", "Leverantör", "Kategori", "Netto", "Moms", "Totalt", "Anteckning"]
    _set_column_widths(ws, [12, 28, 22, 14, 14, 14, 36])
    _write_header_row(ws, row=1, headers=headers)

    total_net = ZERO
    total_vat = ZERO
    total_gross = ZERO

    row = 2
    for receipt in receipts:
        date_str = receipt.date.isoformat() if receipt.date else ""
        net = receipt.net_amount or ZERO
        vat = receipt.vat_amount or ZERO
        gross = receipt.total_amount or ZERO

        ws.cell(row=row, column=1, value=date_str).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=2, value=receipt.vendor or "").alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=3, value=_category_label(receipt)).alignment = BODY_ALIGN_LEFT

        for col, value in [(4, net), (5, vat), (6, gross)]:
            cell = ws.cell(row=row, column=col, value=value)
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = BODY_ALIGN_RIGHT

        ws.cell(row=row, column=7, value=receipt.note or "").alignment = BODY_ALIGN_LEFT

        for col in range(1, 8):
            ws.cell(row=row, column=col).font = BODY_FONT
            ws.cell(row=row, column=col).border = BORDER

        total_net += net
        total_vat += vat
        total_gross += gross
        row += 1

    # Total row at the bottom
    total_row = row
    ws.cell(row=total_row, column=1, value="Summa")
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
    for col, value in [(4, total_net), (5, total_vat), (6, total_gross)]:
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.number_format = CURRENCY_FORMAT
        cell.alignment = BODY_ALIGN_RIGHT
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL

    ws.freeze_panes = "A2"


# --- Sheet 2: Sammanfattning (per-category sums) -----------------------------


def _write_sammanfattning_sheet(wb: Workbook, receipts: list[Receipt]) -> None:
    ws = wb.create_sheet("Sammanfattning")

    headers = ["Kategori", "Antal", "Netto", "Moms", "Totalt"]
    _set_column_widths(ws, [28, 10, 16, 16, 16])
    _write_header_row(ws, row=1, headers=headers)

    # Aggregate per category — preserve insertion order so the output is
    # deterministic and easy to diff between exports
    by_category: OrderedDict[str, dict[str, Decimal | int]] = OrderedDict()
    for receipt in receipts:
        label = _category_label(receipt)
        bucket = by_category.setdefault(
            label, {"count": 0, "net": ZERO, "vat": ZERO, "gross": ZERO}
        )
        bucket["count"] += 1
        bucket["net"] += receipt.net_amount or ZERO
        bucket["vat"] += receipt.vat_amount or ZERO
        bucket["gross"] += receipt.total_amount or ZERO

    # Sort by gross desc so the biggest spend is at top
    sorted_categories = sorted(
        by_category.items(),
        key=lambda item: item[1]["gross"],
        reverse=True,
    )

    total_net = ZERO
    total_vat = ZERO
    total_gross = ZERO
    total_count = 0

    row = 2
    for label, bucket in sorted_categories:
        ws.cell(row=row, column=1, value=label).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=2, value=bucket["count"]).alignment = BODY_ALIGN_RIGHT
        for col, key in [(3, "net"), (4, "vat"), (5, "gross")]:
            cell = ws.cell(row=row, column=col, value=bucket[key])
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = BODY_ALIGN_RIGHT

        for col in range(1, 6):
            ws.cell(row=row, column=col).font = BODY_FONT
            ws.cell(row=row, column=col).border = BORDER

        total_net += bucket["net"]
        total_vat += bucket["vat"]
        total_gross += bucket["gross"]
        total_count += bucket["count"]
        row += 1

    # Total row
    total_row = row
    ws.cell(row=total_row, column=1, value="Summa")
    ws.cell(row=total_row, column=2, value=total_count).alignment = BODY_ALIGN_RIGHT
    for col, value in [(3, total_net), (4, total_vat), (5, total_gross)]:
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.number_format = CURRENCY_FORMAT
        cell.alignment = BODY_ALIGN_RIGHT

    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL

    ws.freeze_panes = "A2"


# --- Public API ---------------------------------------------------------------


def build_excel(receipts: list[Receipt]) -> io.BytesIO:
    """Generate the workbook for a list of receipts.

    Returns an in-memory BytesIO ready to be served as a FileResponse.
    """
    wb = Workbook()
    _write_kvitton_sheet(wb, receipts)
    _write_sammanfattning_sheet(wb, receipts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
